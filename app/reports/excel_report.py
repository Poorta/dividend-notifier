"""
Excel 报表生成模块

使用 openpyxl 生成对标参考表格的行业分组 Excel 报表。

功能:
- 行业分隔行 (深蓝底白字)
- 颜色编码 (股息率/YTD/连续高息年数)
- 冻结标题行 + 行业列
- 红利税脚注
- 列宽自适应
"""

import os
from datetime import datetime
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter

from app.config import settings
from app.utils.colors import (
    EXCEL_RED,
    EXCEL_GREEN,
    EXCEL_HEADER_BG,
    EXCEL_HEADER_FG,
    EXCEL_INDUSTRY_BG,
    EXCEL_FOOTER_BG,
    EXCEL_BORDER_COLOR,
    get_dividend_yield_color,
    get_ytd_return_color,
    get_consecutive_years_color,
)
from app.utils.logger import logger


# === 列定义 ===
COLUMNS = [
    ("行业", "industry", 10),
    ("股票名称", "name", 10),
    ("代码", "code", 8),
    ("市值(亿元)", "market_cap", 12),
    ("连续派现(年)", "consecutive_years", 12),
    ("最新价", "latest_price", 8),
    ("上年每股分红", "annual_dividend", 12),
    ("当前股息率", "dividend_yield", 10),
    ("除权除息日", "ex_dividend_date", 12),
    ("上年末股价", "year_end_price", 12),
    ("今年以来涨跌", "ytd_return", 12),
    ("4%对应股价", "yield_price_4", 11),
    ("5%对应股价", "yield_price_5", 11),
    ("6%对应股价", "yield_price_6", 11),
    ("7%对应股价", "yield_price_7", 11),
    ("8%对应股价", "yield_price_8", 11),
    ("分红明细", "dividend_detail", 18),
]

# 列索引 (0-based)
COL_IDX = {col[1]: i for i, col in enumerate(COLUMNS)}


def _get_report_date() -> str:
    """获取报表日期，默认今天"""
    return datetime.now().strftime("%Y.%m.%d")


def _apply_cell_style(cell, font_color: Optional[str] = None, bold: bool = False):
    """给单元格应用字体颜色和加粗"""
    font_args = {"bold": bold}
    if font_color == "red":
        font_args["color"] = EXCEL_RED
    elif font_color == "green":
        font_args["color"] = EXCEL_GREEN
    cell.font = Font(**font_args)


def generate_excel(
    grouped_stocks: dict[str, pd.DataFrame],
    output_dir: Optional[str] = None,
    report_date: Optional[str] = None,
) -> str:
    """
    生成 Excel 报表。

    Args:
        grouped_stocks: {行业名: DataFrame} 按行业分组+排序后的数据
        output_dir: 输出目录，默认 settings.output.output_dir
        report_date: 报表日期，默认今天

    Returns:
        生成的 xlsx 文件路径
    """
    if output_dir is None:
        output_dir = settings.output.output_dir
    if report_date is None:
        report_date = _get_report_date()

    os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "红利股汇总"

    # === 样式预设 ===
    thin_border = Border(
        left=Side(style="thin", color=EXCEL_BORDER_COLOR),
        right=Side(style="thin", color=EXCEL_BORDER_COLOR),
        top=Side(style="thin", color=EXCEL_BORDER_COLOR),
        bottom=Side(style="thin", color=EXCEL_BORDER_COLOR),
    )
    header_fill = PatternFill(start_color=EXCEL_HEADER_BG, end_color=EXCEL_HEADER_BG, fill_type="solid")
    industry_fill = PatternFill(start_color=EXCEL_INDUSTRY_BG, end_color=EXCEL_INDUSTRY_BG, fill_type="solid")
    footer_fill = PatternFill(start_color=EXCEL_FOOTER_BG, end_color=EXCEL_FOOTER_BG, fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # === Row 1: 标题 ===
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    title_cell = ws.cell(row=1, column=1, value=f"最新养老收息红利股汇总  {report_date}")
    title_cell.font = Font(bold=True, size=14, color="1A3A5C")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # === Row 2: 空行 ===
    ws.row_dimensions[2].height = 8

    # === Row 3: 表头 ===
    header_row = 3
    for col_idx, (col_name, _, col_width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = Font(bold=True, color=EXCEL_HEADER_FG, size=10)
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[header_row].height = 22

    # === 数据行: 按行业分组写入 ===
    current_row = header_row + 1
    total_stocks = 0

    for industry_name, group_df in grouped_stocks.items():
        # --- 行业分隔行 ---
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=len(COLUMNS),
        )
        industry_cell = ws.cell(
            row=current_row, column=1,
            value=f"【{industry_name}】  ({len(group_df)}只)",
        )
        industry_cell.font = Font(bold=True, size=10, color="1A3A5C")
        industry_cell.fill = industry_fill
        industry_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # --- 组内数据行 ---
        for _, stock in group_df.iterrows():
            for col_idx, (col_name, col_key, _) in enumerate(COLUMNS, 1):
                value = stock.get(col_key, "")
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.border = thin_border

                # 数字列居中，文本列左对齐
                if col_key in ("industry", "name", "ex_dividend_date", "dividend_price_impact", "dividend_detail"):
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

                # 颜色编码
                font_color = None
                if col_key == "dividend_yield":
                    font_color = get_dividend_yield_color(value or 0)
                elif col_key == "ytd_return":
                    font_color = get_ytd_return_color(value or 0)
                elif col_key == "consecutive_years":
                    font_color = get_consecutive_years_color(value or 0)

                _apply_cell_style(cell, font_color, bold=False)

            ws.row_dimensions[current_row].height = 20
            current_row += 1
            total_stocks += 1

        # --- 行业组间空行 ---
        current_row += 1

    # === 脚注行 ===
    current_row += 1

    footnotes = [
        "⚠️ 免责声明：本报告仅供信息参考，不构成投资建议。投资有风险，入市需谨慎。",
        "红利税规则：持有期≤1个月 → 红利税20%  |  1个月~1年(含) → 红利税10%  |  >1年 → 免税",
        f"覆盖标的: {total_stocks}只  |  行业数: {len(grouped_stocks)}  |  生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "Powered by AkShare  |  Dividend Notifier (Open Source)",
    ]

    for footnote in footnotes:
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=len(COLUMNS),
        )
        cell = ws.cell(row=current_row, column=1, value=footnote)
        cell.font = Font(size=9, italic=True, color="666666")
        cell.fill = footer_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[current_row].height = 18
        current_row += 1

    # === 冻结窗格 (标题行 + 表头始终可见) ===
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # === 自动筛选 ===
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(COLUMNS))}{current_row - len(footnotes) - 1}"

    # === 保存 ===
    filename = f"dividend_report_{report_date.replace('.', '')}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)

    logger.info(f"Excel 报表已生成: {filepath} ({total_stocks}只股票, {len(grouped_stocks)}个行业)")
    return filepath
